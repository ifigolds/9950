import os
import json
import tempfile
from fastapi.staticfiles import StaticFiles
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import FastAPI, Request, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, Response
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from dotenv import load_dotenv
from aiogram import Bot
from aiogram.types import FSInputFile

from database import (
    init_db,
    seed_data,
    get_all_inventory,
    get_low_stock,
    get_logs,
    add_product,
    use_product,
    delete_product,
    export_backup_data,
    import_backup_data,
    get_connection,
)

load_dotenv()

app = FastAPI()
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

DEVELOPER_PASSWORD = os.getenv("DEVELOPER_PASSWORD", "Dfgnmbxo1")
BOT_TOKEN = os.getenv("BOT_TOKEN", "")


class ProductCreate(BaseModel):
    name: str
    quantity: float
    unit: str
    location: str
    minimum: float = 0
    notes: str = ""


class ProductUse(BaseModel):
    product_id: int
    quantity: float


class LoginRequest(BaseModel):
    mode: str
    password: str | None = None


class BackupSendRequest(BaseModel):
    password: str
    chat_id: int


def build_excel_file() -> BytesIO:
    inventory = get_all_inventory()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = ["ID", "שם מוצר", "כמות", "יחידה", "מיקום", "מינימום", "הערות", "סטטוס"]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in inventory:
        quantity = float(item["quantity"])
        minimum = float(item["minimum"])

        if quantity <= 0:
            status = "נגמר"
        elif quantity <= minimum:
            status = "מחסור"
        else:
            status = "תקין"

        ws.append([
            item["id"],
            item["name"],
            item["quantity"],
            item["unit"],
            item["location"],
            item["minimum"],
            item["notes"],
            status,
        ])

    column_widths = {
        "A": 8,
        "B": 24,
        "C": 12,
        "D": 14,
        "E": 18,
        "F": 12,
        "G": 24,
        "H": 14,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        status_cell = row[7]
        if status_cell.value == "נגמר":
            status_cell.fill = PatternFill(fill_type="solid", fgColor="FEE2E2")
        elif status_cell.value == "מחסור":
            status_cell.fill = PatternFill(fill_type="solid", fgColor="FEF3C7")
        else:
            status_cell.fill = PatternFill(fill_type="solid", fgColor="DCFCE7")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def import_excel_data(file_bytes: bytes) -> None:
    wb = load_workbook(BytesIO(file_bytes))
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Excel ריק")

    data_rows = rows[1:]

    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM inventory")

    for row in data_rows:
        if not row or not row[1]:
            continue

        name = str(row[1] or "").strip()
        quantity = float(row[2] or 0)
        unit = str(row[3] or "").strip()
        location = str(row[4] or "").strip()
        minimum = float(row[5] or 0)
        notes = str(row[6] or "").strip()

        cursor.execute("""
        INSERT INTO inventory (name, quantity, unit, location, minimum, notes)
        VALUES (?, ?, ?, ?, ?, ?)
        """, (name, quantity, unit, location, minimum, notes))

    conn.commit()
    conn.close()


@app.on_event("startup")
def startup():
    init_db()
    seed_data()


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse(request, "index.html")


@app.post("/api/login")
async def login(payload: LoginRequest):
    if payload.mode == "user":
        return {"ok": True, "role": "user"}

    if payload.mode == "developer":
        if payload.password == DEVELOPER_PASSWORD:
            return {"ok": True, "role": "developer"}
        raise HTTPException(status_code=401, detail="סיסמה שגויה")

    raise HTTPException(status_code=400, detail="מצב התחברות לא תקין")


@app.get("/api/inventory")
async def inventory():
    return JSONResponse(get_all_inventory())


@app.get("/api/low-stock")
async def low_stock():
    return JSONResponse(get_low_stock())


@app.get("/api/logs")
async def logs():
    return JSONResponse(get_logs())


@app.post("/api/products")
async def create_product(product: ProductCreate):
    add_product(
        name=product.name,
        quantity=product.quantity,
        unit=product.unit,
        location=product.location,
        minimum=product.minimum,
        notes=product.notes,
    )
    return {"ok": True}


@app.post("/api/use")
async def use_product_api(payload: ProductUse):
    success, message = use_product(payload.product_id, payload.quantity)

    if not success:
        raise HTTPException(status_code=400, detail=message)

    return {"ok": True}


@app.delete("/api/products/{product_id}")
async def delete_product_api(product_id: int):
    success = delete_product(product_id)

    if not success:
        raise HTTPException(status_code=404, detail="המוצר לא נמצא")

    return {"ok": True}


@app.get("/api/backup")
async def download_backup(password: str):
    if password != DEVELOPER_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized")

    data = export_backup_data()
    content = json.dumps(data, ensure_ascii=False, indent=2)

    return Response(
        content=content.encode("utf-8"),
        media_type="application/json; charset=utf-8",
        headers={
            "Content-Disposition": 'attachment; filename="warehouse_backup.json"'
        }
    )


@app.post("/api/backup/import")
async def upload_backup(password: str, file: UploadFile = File(...)):
    if password != DEVELOPER_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized")

    if not file.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="יש להעלות קובץ JSON")

    raw = await file.read()

    try:
        data = json.loads(raw.decode("utf-8"))
        import_backup_data(data)
    except Exception:
        raise HTTPException(status_code=400, detail="קובץ גיבוי לא תקין")

    return {"ok": True}


@app.post("/api/backup/send-to-telegram")
async def send_backup_to_telegram(payload: BackupSendRequest):
    if payload.password != DEVELOPER_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized")

    if not BOT_TOKEN:
        raise HTTPException(status_code=500, detail="BOT_TOKEN חסר בשרת")

    data = export_backup_data()
    content = json.dumps(data, ensure_ascii=False, indent=2)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".json", mode="w", encoding="utf-8") as temp_file:
        temp_file.write(content)
        temp_path = temp_file.name

    try:
        bot = Bot(token=BOT_TOKEN)
        document = FSInputFile(temp_path, filename="warehouse_backup.json")
        await bot.send_document(
            chat_id=payload.chat_id,
            document=document,
            caption="📦 גיבוי מלאי JSON"
        )
        await bot.session.close()
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return {"ok": True}


@app.get("/api/export/excel")
async def export_excel(password: str):
    if password != DEVELOPER_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized")

    output = build_excel_file()

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="warehouse_inventory.xlsx"'
        }
    )


@app.post("/api/export/excel/send-to-telegram")
async def send_excel_to_telegram(payload: BackupSendRequest):
    if payload.password != DEVELOPER_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized")

    if not BOT_TOKEN:
        raise HTTPException(status_code=500, detail="BOT_TOKEN חסר בשרת")

    output = build_excel_file()

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_file.write(output.getvalue())
        temp_path = temp_file.name

    try:
        bot = Bot(token=BOT_TOKEN)
        document = FSInputFile(temp_path, filename="warehouse_inventory.xlsx")
        await bot.send_document(
            chat_id=payload.chat_id,
            document=document,
            caption="📊 דוח מלאי Excel"
        )
        await bot.session.close()
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return {"ok": True}


@app.post("/api/import/excel")
async def import_excel(password: str, file: UploadFile = File(...)):
    if password != DEVELOPER_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized")

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="יש להעלות קובץ Excel")

    raw = await file.read()

    try:
        import_excel_data(raw)
    except Exception:
        raise HTTPException(status_code=400, detail="קובץ Excel לא תקין")

    return {"ok": True}


@app.get("/api/health")
async def health():
    return {"status": "ok"}